import telebot
from telebot.types import ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import Workbook, load_workbook
import os

import os
TOKEN = os.getenv("TOKEN")
ADMIN_ID = 1459646641

bot = telebot.TeleBot(TOKEN)

FILE = "orders.xlsx"
user_data = {}

# =============== EXCEL =================
if not os.path.exists(FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["User ID", "Customer Name", "Files Sent", "Order Count", "Payment Ref", "Status"])
    wb.save(FILE)

def save_order(user_id, name, file_count, payment_ref, status):
    wb = load_workbook(FILE)
    ws = wb.active

    count = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            count = max(count, row[3] + 1)

    ws.append([user_id, name, file_count, count, payment_ref, status])
    wb.save(FILE)

def get_total_files(user_id):
    if not os.path.exists(FILE):
        return 0

    wb = load_workbook(FILE)
    ws = wb.active

    total = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            try:
                total += int(row[2])
            except:
                pass

    return total

# =============== START =================
@bot.message_handler(commands=['start'])
def start(msg):
    markup = ReplyKeyboardMarkup(resize_keyboard=True)
    markup.add(KeyboardButton("Buy File Slot"))
    bot.send_message(msg.chat.id, "Welcome!\nClick below to buy:", reply_markup=markup)

# =============== BUY =================
@bot.message_handler(func=lambda m: m.chat.id != ADMIN_ID and m.text and m.text.lower() == "buy file slot")
def buy(msg):
    bot.send_message(msg.chat.id, "💰 Send your payment reference")
    user_data[msg.chat.id] = {"step": "payment"}

# =============== USER FLOW =================
@bot.message_handler(func=lambda m: m.chat.id != ADMIN_ID, content_types=['text'])
def handle(msg):
    uid = msg.chat.id

    if uid not in user_data:
        return

    step = user_data[uid].get("step")

    username = msg.from_user.username or "NoUsername"
    name = f"{msg.from_user.first_name} {msg.from_user.last_name or ''}".strip()

    # STEP 1: PAYMENT
    if step == "payment":
        user_data[uid]["payment"] = msg.text
        user_data[uid]["name"] = name
        user_data[uid]["username"] = username

        total_files = get_total_files(uid)

        markup = InlineKeyboardMarkup()
        markup.add(
            InlineKeyboardButton("✅ Approve", callback_data=f"approve_{uid}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"reject_{uid}")
        )

        bot.send_message(
            ADMIN_ID,
            f"📥 New Order\n"
            f"User ID: {uid}\n"
            f"Name: {name}\n"
            f"Username: @{username}\n"
            f"Ref: {msg.text}\n"
            f"Total file checked: {total_files}",
            reply_markup=markup
        )

        bot.send_message(uid, "⏳ Waiting approval...")
        user_data[uid]["step"] = "waiting"

    # STEP 2: FILE NAME
    elif step == "approved":
        user_data[uid]["file_name"] = msg.text
        user_data[uid]["file_count"] = 0

        markup = InlineKeyboardMarkup()
        markup.add(
            InlineKeyboardButton("✅ Complete Order", callback_data=f"done_{uid}")
        )

        bot.send_message(
            ADMIN_ID,
            f"📁 File request from {uid}\nFile: {msg.text}\n\nSend files, then click COMPLETE",
            reply_markup=markup
        )

        bot.send_message(uid, "📦 Waiting for files...")
        user_data[uid]["step"] = "file_wait"

# =============== BUTTON HANDLER =================
@bot.callback_query_handler(func=lambda call: True)
def callback(call):
    data = call.data
    action, uid = data.split("_")
    uid = int(uid)

    if uid not in user_data:
        bot.answer_callback_query(call.id, "User not found")
        return

    if action == "approve":
        user_data[uid]["step"] = "approved"
        bot.send_message(uid, "✅ Payment approved!\nSend file name & follow this format- Assignment name: File name")
        bot.send_message(ADMIN_ID, f"✅ Approved {uid}")

    elif action == "reject":
        bot.send_message(uid, "❌ Payment rejected")
        user_data.pop(uid)

    elif action == "done":
        save_order(
            uid,
            f"{user_data[uid]['name']} (@{user_data[uid]['username']})",
            user_data[uid]["file_count"],
            user_data[uid]["payment"],
            "Completed"
        )

        bot.send_message(uid, f"✅ Order completed!\nFiles received: {user_data[uid]['file_count']}")
        bot.send_message(ADMIN_ID, f"✅ Order completed for {uid}")

        user_data.pop(uid)

# =============== SEND FILE =================
@bot.message_handler(content_types=['document'])
def send_file(msg):
    if msg.chat.id != ADMIN_ID:
        return

    for uid in user_data:
        if user_data[uid].get("step") == "file_wait":
            bot.send_document(uid, msg.document.file_id)

            # increase file count
            user_data[uid]["file_count"] += 1
            break

# =============== RUN =================
print("Bot running...")
bot.infinity_polling()